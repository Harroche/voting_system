from openpyxl import workbook, load_workbook  #pip install openpyxl  
import pyrankvote                             #pip install pyrankvote
from pyrankvote import Candidate, Ballot

#loading the excel sheet
wb = load_workbook('voting/example_vote.xlsx',data_only=True)
ws = wb["Chairperson"] #which sheet you want
num_of_seats = int(input("number of seats for the postion: "))
votes =[]
#registers a vote if passed verfication adds to list else voids it 
for row in range(2,ws.max_row+1):
    singular_vote= []
    for col in range(4,ws.max_column+1):
        val=ws.cell(row=row,column=col).value
        if (isinstance(val,int)):
            if (val ==1):
                votes.append(singular_vote)
        else:
            singular_vote.append(val)

#orgainses results in election 
candidates = []
for can in votes[0]:
    candidates.append(Candidate(can))
ballot=[]
for line in votes:  
    ballotArray = []
    for can in line:
        ballotArray.append(Candidate(can))
    ballot.append(Ballot(ranked_candidates=ballotArray))

#prints election results 
election_results = pyrankvote.single_transferable_vote(candidates,ballot,num_of_seats)
print(election_results)